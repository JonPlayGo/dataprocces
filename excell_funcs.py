from openpyxl import load_workbook, Workbook
import datetime


def get_col_data(search_filename, col_name):
    """
    Функция get_col_data принимает имя файла и имя столбца в качестве входных данных, загружает файл
    Excel, ищет столбец с заданным именем и возвращает список значений в этом столбце, начиная со второй
    строки.
    
    :param search_filename: Параметр search_filename — это имя файла Excel (без расширения), в котором
    вы хотите искать данные столбца
    :param col_name: Параметр col_name — это имя столбца, из которого вы хотите получить данные
    :return: список значений из определенного столбца в файле Excel.
    """
    get_col_wb = load_workbook(f"{search_filename}.xlsx")
    get_col_ws = get_col_wb.worksheets[0]
    result_data = []
    for col in range(1, get_col_ws.max_column + 1):
        cell = get_col_ws.cell(row=1, column=col)
        if cell.value == col_name:
            for rw in range(2, get_col_ws.max_row + 1):
                cell2 = get_col_ws.cell(row=rw, column=col)
                result_data.append(cell2.value)
    get_col_wb.save(f"{search_filename}.xlsx")
    get_col_wb.close
    return result_data


def change_col_data(search_filename: str, col_name: str, adding_data: list):
    """
    Функция «change_col_data» принимает имя файла, имя столбца и список данных и добавляет данные в
    новый столбец в файле Excel с заданным именем файла.
    
    :param search_filename: Параметр search_filename — это строка, представляющая имя файла Excel (без
    расширения), который вы хотите изменить
    :type search_filename: str
    :param col_name: Параметр col_name — это строка, представляющая имя столбца, в который будут
    добавлены данные
    :type col_name: str
    :param adding_data: Параметр adding_data — это список значений, которые вы хотите добавить в
    определенный столбец в файле Excel
    :type adding_data: list
    :return: список добавленных данных.
    """
    change_col_wb = load_workbook(f"{search_filename}.xlsx")
    change_col_ws = change_col_wb.worksheets[0]
    result_data = []
    colu = change_col_ws.max_column + 1
    new_column = change_col_ws.cell(row=1, column=change_col_ws.max_column + 1)
    new_column.value = col_name
    rw = 2
    for data in adding_data:
        cell2 = change_col_ws.cell(row=rw, column=colu)
        cell2.value = data
        result_data.append(cell2.value)
        rw = rw + 1
    change_col_wb.save(f"{search_filename}.xlsx")
    change_col_wb.close()
    return result_data


def get_vpr(search_filename, with_col_name, col_name, with_search_data):
    """
    Функция get_vpr принимает имя файла поиска, имя столбца для поиска, данные поиска и имя целевого
    столбца и возвращает соответствующие данные из целевого столбца для строк, соответствующих критериям
    поиска.
    
    :param search_filename: Имя файла Excel, в котором вы хотите выполнить поиск (без расширения файла)
    :param with_col_name: Параметр with_col_name — это имя столбца, в котором вы хотите выполнить поиск
    :param col_name: Параметр col_name — это имя столбца, в котором вы хотите искать данные
    :param with_search_data: Параметр «with_search_data» представляет собой список значений, которые вы
    хотите найти в указанном столбце файла Excel
    :return: список значений из указанного столбца в файле Excel.
    """
    get_col_wb = load_workbook(f"{search_filename}.xlsx")
    get_col_ws = get_col_wb.worksheets[0]
    with_col_data = []
    result_data = []
    for col in range(1, get_col_ws.max_column + 1):
        cell = get_col_ws.cell(row=1, column=col)
        if cell.value == with_col_name:
            for rw in range(2, get_col_ws.max_row + 1):
                cell2 = get_col_ws.cell(row=rw, column=col)
                for search_val in with_search_data:
                    if cell2.value == search_val:
                        with_col_data.append(rw)
    for col2 in range(1, get_col_ws.max_column + 1):
        cell = get_col_ws.cell(row=1, column=col2)
        if cell.value == col_name:
            for with_row in with_col_data:
                cell3 = get_col_ws.cell(row=with_row, column=col2)
                result_data.append(cell3.value)

    get_col_wb.save(f"{search_filename}.xlsx")
    get_col_wb.close
    return result_data


def get_one_vpr(search_filename, with_col_name, col_name, with_search_data):
    """
    Функция get_vpr принимает имя файла поиска, имя столбца для поиска, данные поиска и имя целевого
    столбца и возвращает соответствующие данные из целевого столбца для строк, соответствующих критериям
    поиска.

    :param search_filename: Имя файла Excel, в котором вы хотите выполнить поиск (без расширения файла)
    :param with_col_name: Параметр with_col_name — это имя столбца, в котором вы хотите выполнить поиск
    :param col_name: Параметр col_name — это имя столбца, в котором вы хотите искать данные
    :param with_search_data: Параметр «with_search_data» представляет собой список значений, которые вы
    хотите найти в указанном столбце файла Excel
    :return: список значений из указанного столбца в файле Excel.
    """
    get_col_wb = load_workbook(f"{search_filename}.xlsx")
    get_col_ws = get_col_wb.worksheets[0]
    with_col_data = []
    for col in range(1, get_col_ws.max_column + 1):
        cell = get_col_ws.cell(row=1, column=col)
        if cell.value == with_col_name:
            for rw in range(2, get_col_ws.max_row + 1):
                cell2 = get_col_ws.cell(row=rw, column=col)
                if cell2.value == with_search_data:
                    with_col_data.append(rw)
    for col2 in range(1, get_col_ws.max_column + 1):
        cell = get_col_ws.cell(row=1, column=col2)
        if cell.value == col_name:
            for with_row in with_col_data:
                cell3 = get_col_ws.cell(row=with_row, column=col2)
                get_col_wb.save(f"{search_filename}.xlsx")
                get_col_wb.close
                return cell3.value


def filters(filename, filter_dict: dict, col_name):
    """
    Функция filters принимает имя файла, словарь фильтров и имя столбца в качестве входных данных и
    возвращает список отфильтрованных значений из указанного столбца в файле Excel.
    
    :param filename: Параметр filename — это имя файла Excel (без расширения), который вы хотите
    загрузить и отфильтровать
    :param filter_dict: Параметр filter_dict — это словарь, содержащий фильтры, которые вы хотите
    применить к данным. Ключи словаря представляют имена столбцов, а значения представляют значения
    фильтра для каждого столбца
    :type filter_dict: dict
    :param col_name: Параметр col_name — это имя столбца в файле Excel, по которому вы хотите
    фильтровать
    :return: список значений из указанного столбца («col_name») в файле Excel («filename.xlsx»), которые
    соответствуют критериям фильтрации, указанным в «filter_dict».
    """
    get_col_wb = load_workbook(f"{filename}.xlsx")
    get_col_ws = get_col_wb.worksheets[0]
    result_data = []
    dct = {}
    for filter in filter_dict:
        dct[filter] = []
        for col in range(1, get_col_ws.max_column + 1):
            cell = get_col_ws.cell(row=1, column=col)
            if cell.value in filter:
                for rw in range(2, get_col_ws.max_row + 1):
                    cell2 = get_col_ws.cell(row=rw, column=col)
                    if filter == "Дата транзакции":
                        current_date = datetime.datetime.today()
                        future_date = current_date - datetime.timedelta(days=365)
                        month = future_date.month
                        day = 1
                        rounded_date = datetime.datetime(future_date.year, month, day)
                        if cell2.value >= rounded_date:
                            dct["Дата транзакции"].append(rw)
                    elif type(filter_dict[filter]) == type([]):
                        for filter_value in filter_dict[filter]:
                            if cell2.value == filter_value:
                                dct[filter].append(rw)
                                break
                    else:
                        if cell2.value == filter_dict[filter]:
                            dct[filter].append(rw)

    for fil in dct:
        result_data.append(dct[fil])
    common_values = set(result_data[0])
    for list in result_data[1:]:
        common_values = common_values & set(list)
    result_dataz = []
    for col in range(1, get_col_ws.max_column + 1):
        cell = get_col_ws.cell(row=1, column=col)
        if cell.value == col_name:
            for rw in common_values:
                cell2 = get_col_ws.cell(row=rw, column=col)
                result_dataz.append(cell2.value)

    get_col_wb.save(f"{filename}.xlsx")
    get_col_wb.close
    return result_dataz


def new_file(filename):
    """
    Функция создает новый файл Excel с заданным именем.
    
    :param filename: Параметр filename — это строка, представляющая имя файла, который вы хотите создать
    """
    calculate_workbook = Workbook()
    calculate_workbook.save(filename)
    calculate_workbook.close()


def clean_file(filename):
    get_col_wb = load_workbook(f"{filename}.xlsx")
    get_col_ws = get_col_wb.worksheets[0]

    for merge in list(get_col_ws.merged_cells.ranges):
        get_col_ws.unmerge_cells(range_string=str(merge))

    for col in range(1, get_col_ws.max_column + 1):
        for rw in range(1, get_col_ws.max_row + 1):
            cell1 = get_col_ws.cell(row=rw, column=col)
            if not cell1.value == None:
                if col == 1:
                    cell2 = get_col_ws.cell(row=rw, column=col)
                else:
                    cell2 = get_col_ws.cell(row=rw, column=col - 1)
                if cell2.value == None:
                    minus = 1
                    while cell2.value == None:
                        get_col_ws.move_range(cell1.coordinate, rows=0, cols=-1)
                        minus = minus + 1
                        cell2 = get_col_ws.cell(row=rw, column=col - minus)

    for col in range(1, get_col_ws.max_column + 1):
        for rw in range(1, get_col_ws.max_row + 1):
            cell = get_col_ws.cell(row=rw, column=col)
            if cell.value == None:
                get_col_ws.delete_rows(rw)

    for col in range(1, get_col_ws.max_column + 1):
        for rw in range(1, get_col_ws.max_row + 1):
            cell = get_col_ws.cell(row=rw, column=col)
            if type(cell.value) == str:
                cell.value = cell.value.strip()
    get_col_wb.save("4. Данные по списаниям_2.xlsx")
    get_col_wb.close
